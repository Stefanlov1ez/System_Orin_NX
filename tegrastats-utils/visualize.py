# coding: utf8
import argparse
import re
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference

tegrastats_version = 3.2

class Status:
    def __init__(self, num_cpus=6):
        self.cpu_freq = [[] for _ in range(num_cpus)]
        self.cpu_utilization = [[] for _ in range(num_cpus)]
        self.ram = []
        self.emc = []
        self.gpu = []

def filter_content_from_raw_log(start_time, end_time, file_path):
    with open(file_path, 'r') as fin:
        content = fin.read()
    
    if not start_time or not end_time:
        time_stamps = re.findall(r'(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})', content)
        if time_stamps:
            if not start_time:
                start_time = time_stamps[0]
            if not end_time:
                end_time = time_stamps[-1]
        else:
            print("无法从日志中提取时间戳。")
            return Status()

    print("start time:\t%s" % start_time)
    print("end time:\t%s" % end_time)

    reg = re.compile(r'%s[\s\S]+%s' % (start_time, end_time))
    if tegrastats_version >= 3.2:
        reg2 = re.compile(r'RAM (\d+).+?CPU \[(.+?)\] EMC_FREQ (\d+)%@.+?GR3D_FREQ (\d+)%')
    else:
        reg2 = re.compile(r'RAM (\d+).+?cpu \[(.+?)\] EMC (\d+)%@.+?GR3D (\d+)%')
    
    reg_cpu = re.compile(r'(\d+)\%@(\d+)')
    result = reg.findall(content)
    if not result:
        print("没有符合条件的输出-1")
        return Status()

    # 动态确定CPU数量
    sample_line = result[0]
    cpu_data_match = reg2.findall(sample_line)
    if not cpu_data_match:
        print("没有符合条件的输出-2")
        return Status()
    
    num_cpus = len(cpu_data_match[0][1].split(','))
    status = Status(num_cpus=num_cpus)
    
    for line in cpu_data_match:
        status.ram.append(int(line[0]))
        temp_cpus = line[1].split(',')
        cpu_util_sum = 0
        for i, a_cpu in enumerate(temp_cpus):
            cpu_result = reg_cpu.findall(a_cpu)
            if cpu_result:
                cpu_result = cpu_result[0]
                cpu_util_sum += int(cpu_result[0])
                status.cpu_freq[i].append(int(cpu_result[1]))
            else:
                cpu_util_sum += 0
                status.cpu_freq[i].append(0)
        # 计算平均 CPU 利用率
        cpu_util_avg = cpu_util_sum / num_cpus
        status.cpu_utilization[0].append(cpu_util_avg)
        status.emc.append(int(line[2]))
        status.gpu.append(int(line[3]))

    return status

def write_status_to_xls(status, xls_file='./log.xls'):
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'cpu'
    sheet.cell(row=1, column=1, value='cpu average frequency')
    sheet.cell(row=2, column=1, value='cpu average utilization')
    for i in range(0, len(status.cpu_utilization[0])):
        sheet.cell(row=i + 3, column=1, value=status.cpu_utilization[0][i])
    
    base_offset_ram = 2
    sheet.cell(row=2, column=base_offset_ram + 1, value='ram')
    for i in range(0, len(status.ram)):
        sheet.cell(row=i + 3, column=base_offset_ram + 1, value=status.ram[i])
    
    base_offset_emc = base_offset_ram + 1
    sheet.cell(row=2, column=base_offset_emc + 1, value='emc')
    for i in range(0, len(status.emc)):
        sheet.cell(row=i + 3, column=base_offset_emc + 1, value=status.emc[i])
    
    base_offset_gpu = base_offset_emc + 1
    sheet.cell(row=2, column=base_offset_gpu + 1, value='gpu')
    for i in range(0, len(status.gpu)):
        sheet.cell(row=i + 3, column=base_offset_gpu + 1, value=status.gpu[i])

    chart = LineChart()
    chart.y_axis.title = 'Rate'
    data = Reference(sheet, min_col=1, min_row=2, max_row=len(status.cpu_utilization[0]) + 2, max_col=1)
    data_gpu = Reference(sheet, min_col=base_offset_gpu + 1, min_row=2, max_row=len(status.gpu) + 2, max_col=base_offset_gpu + 1)
    chart.add_data(data, titles_from_data=True)
    chart.add_data(data_gpu, titles_from_data=True)
    sheet.add_chart(chart, 'A%d' % (len(status.gpu) + 5))

    wb.save(xls_file)
    print("Excel file saved to %s" % xls_file)

def plot_status(status, image_file='./status_plot.png'):
    plt.figure(figsize=(10, 6))

    # Plot average CPU utilization
    plt.plot(status.cpu_utilization[0], label='CPU Average Utilization')
    
    # Plot GPU utilization
    plt.plot(status.gpu, label='GPU Utilization', color='red', linewidth=2)

    plt.title('Average CPU and GPU Utilization')
    plt.xlabel('Time (samples)')
    plt.ylabel('Utilization (%)')
    plt.legend(loc='upper right')
    plt.grid(True)
    plt.savefig(image_file, format='png')
    plt.close()
    print(f"Plot saved to {image_file}")

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='用来提取一段时间内tegra cpu gpu利用率等信息的工具')
    parser.add_argument('-s', '--start', metavar='开始的时间戳', required=False, dest='start_time', action='store')
    parser.add_argument('-e', '--end', metavar='结束的时间', required=False, dest='end_time', action='store')
    parser.add_argument('-i', '--input', metavar='输入的log文件', required=False, dest='log_file_path', action='store', default='./a.log')
    parser.add_argument('-o', '--output', metavar='输出的excel文件路径', required=False, dest='xls_file_path', action='store', default='./freq.xls')
    parser.add_argument('--plot', metavar='输出的图像文件路径', required=False, dest='image_file_path', action='store', default='./status_plot.png')
    args = parser.parse_args()

    log_file = args.log_file_path
    start_time = args.start_time
    end_time = args.end_time
    xls_file = args.xls_file_path
    image_file = args.image_file_path

    status = filter_content_from_raw_log(start_time, end_time, log_file)
    write_status_to_xls(status, xls_file)
    plot_status(status, image_file)
